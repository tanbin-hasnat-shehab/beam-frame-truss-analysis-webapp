import streamlit as st
import codecs
import streamlit.components.v1 as stc
from anastruct import SystemElements
import math
import matplotlib.pyplot as plt
from PIL import Image
from openpyxl import load_workbook,Workbook
import json
wb=Workbook()
  
wb.create_sheet("DISPLACEMENTS")
wb.create_sheet('BENDING_MOMENT')
wb.create_sheet('AXIAL_FORCE')
wb.create_sheet('SHEAR_FORCE')
wb.create_sheet('REACTIONS')
wb.save('data.xlsx')


st.set_page_config(layout="wide")

js_str=st.text_input('paste here')
aaa=st.slider("zoom in out results", 0, 300, value=30)
if st.button("Run"):
	
	#js_string = cookie_manager.get(cookie)
	#jtopy=json.dumps(js_str)
	value=json.loads(js_str)

	
	

	#print(value)
		



	ss = SystemElements()

	new_data=[]
	for i in range(len(value)):
		if value[i] not  in new_data:
			new_data.append(value[i])
	for data in new_data:
		if data['type']=='type_of_st':
			structure_type=data['value']
		if data['type']=='grid_spacing':
			grid=data['value']
	for data in new_data:
		if data['type']=='line':
			data['x1']=data['x1']/grid
			data['y1']=data['y1']/grid
			data['x2']=data['x2']/grid
			data['y2']=data['y2']/grid
		if data['type']=='fixed_support':
			data['x']=data['x']/grid
			data['y']=data['y']/grid
			
		if data['type']=='hinged_support':
			data['x']=data['x']/grid
			data['y']=data['y']/grid
			
		if data['type']=='point_load':
			data['pos_x']=data['pos_x']/grid
			data['pos_y']=data['pos_y']/grid


	my_points_x=[]
	my_points_y=[]

	for data in new_data:
		if data['type']=='line':
			my_points_x.append(data['x1'])
			my_points_y.append(data['y1'])
			my_points_x.append(data['x2'])
			my_points_y.append(data['y2'])
		if data['type']=='fixed_support':
			my_points_x.append(data['x'])
			my_points_y.append(data['y'])
			
		if data['type']=='hinged_support':
			my_points_x.append(data['x'])
			my_points_y.append(data['y'])
			
		if data['type']=='point_load':
			my_points_x.append(data['pos_x'])
			my_points_y.append(data['pos_y'])
			

				

		
	print(new_data)
	
	no_of_member=0
	members=[]
	no_of_load=0
	loads=[]
	no_of_fixed=0
	fixed=[]
	no_of_hinged=0
	hinged=[]

	for i in range(len(new_data)):
		if new_data[i]['type']=='line':
			no_of_member+=1
			members.append(new_data[i])
		if new_data[i]['type']=='point_load':
			no_of_load+=1
			loads.append(new_data[i])
		if new_data[i]['type']=='fixed_support':
			no_of_fixed+=1
			fixed.append(new_data[i])
		if new_data[i]['type']=='hinged_support':
			no_of_hinged+=1
			hinged.append(new_data[i])


	if structure_type=='truss':
		for i in range(no_of_member):
			ss.add_element(location=[[members[i]['x1'], members[i]['y1']]   , [members[i]['x2'], members[i]['y2']] ],EA=5000*members[i]['EA'])    
	

	if structure_type=='frame' or structure_type=='beam':
		for i in range(no_of_member):
			ss.add_element(location=[[members[i]['x1'], members[i]['y1']]   , [members[i]['x2'], members[i]['y2']] ],EA=5000*members[i]['EA'])    
			ss.q_load(q=-1*members[i]['W'], element_id=ss.id_last_element, direction='element')
		

	for i in range(no_of_load):
		ss.point_load(Fx=loads[i]['load'],rotation=loads[i]['rotation'], node_id=ss.find_node_id([loads[i]['pos_x'],loads[i]['pos_y']]))
		





	for i in range(no_of_fixed):
		ss.add_support_fixed(node_id=ss.find_node_id([fixed[i]['x'],fixed[i]['y']]))






	for i in range(no_of_hinged):
		
		ss.add_support_hinged(node_id=ss.find_node_id([hinged[i]['x'],hinged[i]['y']]))
	ss.solve()

	wb=load_workbook('data.xlsx')
	dis=wb['DISPLACEMENTS']
	ben=wb['BENDING_MOMENT']
	axi=wb['AXIAL_FORCE']
	shr=wb['SHEAR_FORCE']
	rea=wb['REACTIONS']	
	

	ss.show_structure()
	plt.xlim([min(my_points_x)-aaa,max(my_points_x)+aaa])
	plt.ylim([min(my_points_y)-aaa,max(my_points_y)+aaa])
	plt.title('STRUCTURE')
	plt.savefig('my-figure1.png')
	image1 = Image.open('my-figure1.png')
	st.image(image1)
	
	ss.show_reaction_force()
	plt.xlim([min(my_points_x)-aaa,max(my_points_x)+aaa])
	plt.ylim([min(my_points_y)-aaa,max(my_points_y)+aaa])
	plt.title('REACTIONS')
	plt.savefig('my-figure2.png')
	image2 = Image.open('my-figure2.png')
	st.image(image2)
	'''
	m=ss.show_reaction_force(show=False,values_only=True)
	for i in range(0,len(m[0])):
		rea.cell(row=i+1,column=1).value=m[0][i]
		rea.cell(row=i+1,column=2).value=m[1][i]
	'''


	ss.show_displacement()
	plt.xlim([min(my_points_x)-aaa,max(my_points_x)+aaa])
	plt.ylim([min(my_points_y)-aaa,max(my_points_y)+aaa])
	plt.title('DISPLACEMENTS')
	plt.savefig('my-figure22.png')
	image22 = Image.open('my-figure22.png')
	st.image(image22)
	m=ss.show_displacement(show=False,values_only=True)
	for i in range(0,len(m[0])):
		dis.cell(row=i+1,column=1).value=m[0][i]
		dis.cell(row=i+1,column=2).value=m[1][i]
	
	



	ss.show_axial_force()
	plt.xlim([min(my_points_x)-aaa,max(my_points_x)+aaa])
	plt.ylim([min(my_points_y)-aaa,max(my_points_y)+aaa])
	plt.title('AXIAL FORCE DIAGRAM')
	plt.savefig('my-figure3.png')
	image3 = Image.open('my-figure3.png')
	st.image(image3)
	m=ss.show_axial_force(show=False,values_only=True)
	for i in range(0,len(m[0])):
		axi.cell(row=i+1,column=1).value=m[0][i]
		axi.cell(row=i+1,column=2).value=m[1][i]

	if structure_type=='beam' or structure_type=='frame':
		ss.show_bending_moment()
		plt.xlim([min(my_points_x)-aaa,max(my_points_x)+aaa])
		plt.ylim([min(my_points_y)-aaa,max(my_points_y)+aaa])
		plt.title('BENDING MOMENT DIAGRAM')
		plt.savefig('my-figure4.png')
		image4 = Image.open('my-figure4.png')
		st.image(image4)
		m=ss.show_bending_moment(show=False,values_only=True)
		for i in range(0,len(m[0])):
			ben.cell(row=i+1,column=1).value=m[0][i]
			ben.cell(row=i+1,column=2).value=m[1][i]

		ss.show_shear_force()
		plt.xlim([min(my_points_x)-aaa,max(my_points_x)+aaa])
		plt.ylim([min(my_points_y)-aaa,max(my_points_y)+aaa])
		plt.title('SHEAR FORCE DIAGRAM')
		plt.savefig('my-figure5.png')
		image5 = Image.open('my-figure5.png')
		st.image(image5)
		m=ss.show_shear_force(show=False,values_only=True)
		for i in range(0,len(m[0])):
			shr.cell(row=i+1,column=1).value=m[0][i]
			shr.cell(row=i+1,column=2).value=m[1][i]
	
	wb.save('data.xlsx')

def my_html(html_file,width=2000,height=2000):
	html_file=codecs.open(html_file,'r')
	page=html_file.read()
	stc.html(page,width=width,height=height,scrolling=False)


my_html('a.html')






